from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import openpyxl
import time

exported_xlsx = "C:\\Users\\TAvram\\Desktop\\Lines_Full_validation_tests\\Exported_HPP-DPEO.xlsx"
wb = load_workbook(exported_xlsx)
#save_it = wb.save(exported_xlsx)

ws = wb.active

row_1 = ws['1']
col_a = ws['A']
header = []
column = []



for cell in row_1:
	if cell.value is not None:
		header.append(cell.value)

for cell in col_a:
	if cell.value is not None:
		column.append(cell.value)

for c in range(1,len(header)+1):
	for r in range(1,len(column)+1):
		if ws[f'{get_column_letter(c)}{r}'].value == " ":
			ws[f'{get_column_letter(c)}{r}'].value = 0

# Copy active worksheet
'''def copy_active_worksheet (worksheet_no,test_case_name):
	#worksheet_no = wb.copy_worksheet (wb.active)
	worksheet_no.title = test_case_name
'''

# Filter lower than
def filter_lower_than(work_sheet,signal,value_to_filter):
	for row in range(2,len(column)):
		if signal in header:
			if work_sheet[f'{get_column_letter(header.index(signal)+1)}{row}'].value < value_to_filter:
				work_sheet.row_dimensions[row].hidden = True

# Filter lower than or lower than
def filter_lower_than_or(work_sheet,signal_1,signal_2,value_to_filter):
	for row in range(2,len(column)):
		if signal_1 in header:
			if work_sheet[f'{get_column_letter(header.index(signal_1)+1)}{row}'].value < value_to_filter:
				work_sheet.row_dimensions[row].hidden = True
		elif signal_2 in header:
			if work_sheet[f'{get_column_letter(header.index(signal_2)+1)}{row}'].value < value_to_filter:
				work_sheet.row_dimensions[row].hidden = True

# Filter higher than
def filter_higher_than(work_sheet,signal,value_to_filter):
	for row in range(2,len(column)):
		if signal in header:
			if work_sheet[f'{get_column_letter(header.index(signal)+1)}{row}'].value > value_to_filter:
				work_sheet.row_dimensions[row].hidden = True


# Filter higher than or higher than
def filter_higher_than_or(work_sheet,signal_1,signa_2,value_to_filter):
	for row in range(2,len(column)):
		if signal_1 in header:
			if work_sheet[f'{get_column_letter(header.index(signal_1)+1)}{row}'].value > value_to_filter:
				work_sheet.row_dimensions[row].hidden = True
		elif signal_2 in header:
			if work_sheet[f'{get_column_letter(header.index(signal_2)+1)}{row}'].value > value_to_filter:
				work_sheet.row_dimensions[row].hidden = True

# Filter equal to
def filter_equal_to(work_sheet,signal,value_to_filter):
	for row in range(2,len(column)):
		if signal in header:
			if work_sheet[f'{get_column_letter(header.index(signal)+1)}{row}'].value == value_to_filter:
				work_sheet.row_dimensions[row].hidden = True

# Filter different to
def filter_different_to(work_sheet,signal,value_to_filter):
	for row in range(2,len(column)):
		if signal in header:
			if work_sheet[f'{get_column_letter(header.index(signal)+1)}{row}'].value != value_to_filter:
				work_sheet.row_dimensions[row].hidden = True

# Filter different to or different to
def filter_different_to_or(work_sheet,signal_1,signal_2,value_to_filter):
	for row in range(2,len(column)):
		if signal_1 in header:
			if work_sheet[f'{get_column_letter(header.index(signal_1)+1)}{row}'].value != value_to_filter:
				work_sheet.row_dimensions[row].hidden = True
		elif signal_2 in header:
			if work_sheet[f'{get_column_letter(header.index(signal_2)+1)}{row}'].value != value_to_filter:
				work_sheet.row_dimensions[row].hidden = True

# Filter not same sign
def filter_not_same_sign(work_sheet,signal_1,signal_2):
	for row in range(2,len(column)):
		if signal_1 in header:
			if work_sheet[f'{get_column_letter(header.index(signal_1)+1)}{row}'].value*work_sheet[f'{get_column_letter(header.index(signal_2)+1)}{row}'].value < 0:
				work_sheet.row_dimensions[row].hidden = True