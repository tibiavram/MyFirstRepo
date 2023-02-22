from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
#wb = Workbook()

# load existing spreadsheet
wb = load_workbook('Hello.xlsx')

# Create an active worksheet
ws = wb.active

# Set variable
#name = ws["A3"].value
#color = ws["B3"].value

# Print something from our spreadsheet
#print (f'{name} : {color}')

# Grab whole column

#For loop

# Grab a range
range = ws['A2' : 'A10']
print (range)

names = []
# Loop
for cell in range:
	for x in cell:
		names += x.value
print (names)

column_a = ws['A']
print (column_a)

first_column = []
for cell in column_a:
	first_column += cell.value
print(first_column)